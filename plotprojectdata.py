use_csv=False
if use_csv:
    import csv
else:
    from openpyxl import load_workbook
import argparse
import datetime
import numpy as np
import matplotlib.pyplot as plt 
from scipy.optimize import curve_fit

delimiter = ";"

def linear_func(x, a, b):
    return a + b * x

def getbillingprice(wbs):
    data_dict={}

    found=0
    for i in range(1,wbs.max_column+1):
        string=wbs.cell(1,i).value
        if string == "Billing Price, Reg.":
            billing_index=i
            found+=1
        if string == "Date":
            date_index=i
            found+=1
        if found==2:
            break
    for i in range(2,wbs.max_row+1):
        datestr=wbs.cell(i,date_index).value
        date = datetime.datetime.strptime(datestr.replace("=Date(","").replace(")",""), "%Y,%m,%d")
        datestr = str(date.day)+"."+str(date.month)+"."+str(date.year)

        valuestr=wbs.cell(i,billing_index).value
        value = int(valuestr.replace("=", "").split(".", 1)[0])

        if datestr in data_dict:
            data_dict[datestr] += value
        else:
            data_dict[datestr] = value
    return data_dict


def getbillingprice_csv(fn):
    data={}
    'Date'
    with open(fn) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=delimiter)
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                billing_index=row.index('Billing Price')
                date_index=row.index('Entry Date')
            else:
                if row[date_index] in data:
                    data[row[date_index]] += int(row[billing_index].replace(" ", "").split(",", 1)[0])
                else:
                    data[row[date_index]] = int(row[billing_index].replace(" ", "").split(",", 1)[0])
            line_count += 1
    return data

def getemployees(wbs):
    data_dict={}

    found=0
    for i in range(1,wbs.max_column+1):
        string=wbs.cell(1,i).value
        if string == "Employee No.":
            emplno_index=i
            found+=1
        if string == "Employee Name":
            emplname_index=i
            found+=1
        if found==2:
            break
    for i in range(2,wbs.max_row+1):
        number=wbs.cell(i,emplno_index).value
        name=wbs.cell(i,emplname_index).value
        if name=='':
            name="other"
        data_dict[number] = name
    return data_dict

def getemployees_csv(fn):
    data={}
    with open(fn) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=delimiter)
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                emplno_index=row.index('Empl. No.')
                emplname_index=row.index('Empl. Name')
            else:
                name=str(row[emplname_index])
                if name=='':
                    name="other"
                data[row[emplno_index]] = name
            line_count += 1
    return data

def getemployeesbillingprice(employees_by_number, wbs):
    data_dict={}

    for key in employees_by_number:
        data_dict[key]={}

    found=0
    for i in range(1,wbs.max_column+1):
        string=wbs.cell(1,i).value
        if string == "Employee No.":
            emplno_index=i
            found+=1
        if string == "Billing Price, Reg.":
            billing_index=i
            found+=1
        if string == "Date":
            date_index=i
            found+=1
        if found==3:
            break
    for i in range(2,wbs.max_row+1):
        datestr=wbs.cell(i,date_index).value
        date = datetime.datetime.strptime(datestr.replace("=Date(","").replace(")",""), "%Y,%m,%d")
        datestr = str(date.day)+"."+str(date.month)+"."+str(date.year)

        valuestr=wbs.cell(i,billing_index).value
        value = int(valuestr.replace("=", "").split(".", 1)[0])

        number=wbs.cell(i,emplno_index).value

        if datestr in data_dict[number]:
            data_dict[number][datestr] += value
        else:
            data_dict[number][datestr] = value
    return data_dict

def getemployeesbillingprice_csv(employees_by_number, fn):
    data={}
    for key in employees_by_number:
        data[key]={}
    with open(fn) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=delimiter)
        line_count = 0
        for row in csv_reader:
            if line_count == 0:
                emplno_index=row.index('Empl. No.')
                billing_index=row.index('Billing Price')
                date_index=row.index('Entry Date')
            else:
                if row[date_index] in data[row[emplno_index]]:
                    data[row[emplno_index]][row[date_index]] += int(row[billing_index].replace(" ", "").split(",", 1)[0])
                else:
                    data[row[emplno_index]][row[date_index]] = int(row[billing_index].replace(" ", "").split(",", 1)[0])
            line_count += 1
    return data

def str2bool(v):
    if isinstance(v, bool):
       return v
    if v.lower() in ('yes', 'true', 'True', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'False', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description='Parse and plot data')
    parser.add_argument('--filename', metavar='filename', required=True, type=str, help='name of cvs file')
    parser.add_argument('--totalbudget', metavar='totalbudget', required=False, type=int, help='total budget in KNOK')
    parser.add_argument('--regressionON', metavar='regressionON', type=str2bool, nargs='?', const=True, default=True, help='plot regression')
    args = parser.parse_args()

    if use_csv:
        billings_by_day = getbillingprice_csv(args.filename)
        ### there might be two people with the exact name, we need to use the Empl. No.
        employees_by_number = getemployees_csv(args.filename)
        billings_by_employees_by_day = getemployeesbillingprice_csv(employees_by_number, args.filename)
    else:
        wb = load_workbook(filename=args.filename)
        wbs = wb[wb.sheetnames[0]]

        print("Reading billing prices...", end=" ", flush=True)
        billings_by_day = getbillingprice(wbs)
        print("done.")
        print("Reading employees...", end=" ", flush=True)
        ### there might be two people with the exact name, we need to use the Empl. No.
        employees_by_number = getemployees(wbs)
        print("done.")
        print("Reading billings by employees...", end=" ", flush=True)
        billings_by_employees_by_day = getemployeesbillingprice(employees_by_number, wbs)
        print("done.")
        wb.close()

### if the project is from a past year, set month to 12 and week to number of weeks that year
    today = datetime.datetime.today()
    this_week=today.isocalendar()[1]
    this_month=today.month
    this_year=today.year

    datestr = list(billings_by_day.keys())[0]
    file_date = datetime.datetime.strptime(datestr, "%d.%m.%Y")

    ### any date from the file will suffice
    if this_year > file_date.year:
        year = file_date.year
        month = 12
        week = datetime.date(file_date.year, 12, 29).isocalendar()[1]
    else:
        year = this_year
        month = this_month
        week = this_week

    num_weeks = datetime.date(year, 12, 29).isocalendar()[1]


    billings_by_employees_by_year = {}
    billings_by_employees_by_month = {}
    billings_by_employees_by_week = {}
    for employeenr, billings in billings_by_employees_by_day.items():
        billings_by_employees_by_year[employeenr] = 0
        billings_by_employees_by_month[employeenr] = np.zeros(12)
        billings_by_employees_by_week[employeenr] = np.zeros(num_weeks)
        for billingday, value in billings.items():
            billings_by_employees_by_year[employeenr] += value
            date = datetime.datetime.strptime(billingday, "%d.%m.%Y")
            billings_by_employees_by_month[employeenr][date.month-1] += value
            billings_by_employees_by_week[employeenr][date.isocalendar()[1]-1] += value

    billings_by_month = np.zeros(12)
    billings_by_week = np.zeros(num_weeks)
    for billingday, billings in billings_by_day.items():
        date = datetime.datetime.strptime(billingday, "%d.%m.%Y")
        billings_by_month[date.month-1] += billings
        billings_by_week[date.isocalendar()[1]-1] += billings

    cumsum_billings_by_month=np.cumsum(billings_by_month)
    cumsum_billings_by_month[month:] = 0

    cumsum_billings_by_week=np.cumsum(billings_by_week)
    cumsum_billings_by_week[week:] = 0


    labels_month=('Jan','Feb','Mar','Apr','Mai','Jun','Jul','Aug','Sep','Oct','Nov','Des')

    plt.rcParams["figure.figsize"] = 12,8
    plt.rcParams["axes.titlesize"] = 24
    plt.rcParams["axes.labelsize"] = 20
    plt.rcParams["lines.linewidth"] = 3
    plt.rcParams["lines.markersize"] = 10
    plt.rcParams["xtick.labelsize"] = 16
    plt.rcParams["ytick.labelsize"] = 16
    plt.style.use('bmh')

#### actuals per month
    text='Actuals per month'
    print("generating figure:", text)
    fig, ax = plt.subplots()
    ax.bar(np.linspace(1,12,12),billings_by_month/1000)
    ax.set_ylabel('KNOK')
    ax.set_xticks(np.arange(1,13))
    ax.set_xticklabels(labels_month)
    if args.totalbudget:
        ax.plot([1,12], [args.totalbudget/12, args.totalbudget/12],':k',label='average budget/month')
        ax.legend()
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

    ### per employee
    text='Actuals per month per employee'
    print("generating figure:", text)
    fig, (ax,lax) = plt.subplots(ncols=2, gridspec_kw={"width_ratios":[4,1]})
    bot = np.zeros(12)
    for a, b in billings_by_employees_by_month.items():
        ax.bar(np.linspace(1,12,12),b/1000, bottom=bot/1000, label=employees_by_number[a])
        bot+=b
    ax.set_ylabel('KNOK')
    ax.set_xticks(np.arange(1,13))
    ax.set_xticklabels(labels_month)
    if args.totalbudget:
        ax.plot([1,12], [args.totalbudget/12, args.totalbudget/12],':k',label='average budget/month')
    h,l = ax.get_legend_handles_labels()
    lax.legend(h,l, borderaxespad=0)
    lax.axis("off")
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")



### actuals accumulated per month
    text='Actuals accumulated per month'
    print("generating figure:", text)
    fig, ax = plt.subplots()
    ax.bar(np.linspace(1,12,12),cumsum_billings_by_month/1000)
    ax.set_ylabel('KNOK')
    ax.set_xticks(np.arange(1,13))
    ax.set_xticklabels(labels_month)
    if args.totalbudget:
        ax.plot([1,12], [args.totalbudget, args.totalbudget],'-k',label='total budget')

    if args.regressionON and month>2:
        ydata = cumsum_billings_by_month[:month]
        xdata = np.arange(1,month+1)
        popt, pcov = curve_fit(linear_func, xdata, ydata)
        ax.plot(np.arange(1,13),linear_func(np.arange(1,13), popt[0], popt[1])/1000,':k',label='linear regression')
        ax.plot(12,linear_func(12, popt[0], popt[1])/1000,'ko')
        ax.legend()
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

    ### per employee
    text='Actuals accumulated per month per employee'
    print("generating figure:", text)
    fig, (ax,lax) = plt.subplots(ncols=2, gridspec_kw={"width_ratios":[4,1]})
    bot = np.zeros(12)
    for a, b in billings_by_employees_by_month.items():
        bc = np.cumsum(b)
        bc[month:] = 0
        ax.bar(np.linspace(1,12,12),bc/1000, bottom=bot/1000, label=employees_by_number[a])
        bot+=bc
    ax.set_ylabel('KNOK')
    ax.set_xticks(np.arange(1,13))
    ax.set_xticklabels(labels_month)
    if args.totalbudget:
        ax.plot([1,12], [args.totalbudget, args.totalbudget],'-k',label='total budget')

    if args.regressionON and month>2:
        ax.plot(np.arange(1,13),linear_func(np.arange(1,13), popt[0], popt[1])/1000,':k',label='linear regression')
        ax.plot(12,linear_func(12, popt[0], popt[1])/1000,'ko')
    h,l = ax.get_legend_handles_labels()
    lax.legend(h,l, borderaxespad=0)
    lax.axis("off")
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

#### actuals per week
    text='Actuals per week'
    print("generating figure:", text)
    fig, ax = plt.subplots()
    ax.bar(np.linspace(1,num_weeks,num_weeks),billings_by_week/1000)
    ax.set_ylabel('KNOK')
    if args.totalbudget:
        ax.plot([1,num_weeks], [args.totalbudget/num_weeks, args.totalbudget/num_weeks],':k',label='average budget/week')
        ax.legend()
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

    ### per employee
    text='Actuals per week per employee'
    print("generating figure:", text)
    fig, (ax,lax) = plt.subplots(ncols=2, gridspec_kw={"width_ratios":[4,1]})
    bot = np.zeros(num_weeks)
    for a, b in billings_by_employees_by_week.items():
        ax.bar(np.linspace(1,num_weeks,num_weeks),b/1000, bottom=bot/1000, label=employees_by_number[a])
        bot+=b
    ax.set_ylabel('KNOK')
    if args.totalbudget:
        ax.plot([1,num_weeks], [args.totalbudget/num_weeks, args.totalbudget/num_weeks],':k',label='average budget/week')
    h,l = ax.get_legend_handles_labels()
    lax.legend(h,l, borderaxespad=0)
    lax.axis("off")
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")


### actuals accumulated per week
    text='Actuals accumulated per week'
    print("generating figure:", text)
    fig, ax = plt.subplots()
    ax.bar(np.linspace(1,num_weeks,num_weeks),cumsum_billings_by_week/1000)
    ax.set_ylabel('KNOK')
    if args.totalbudget:
        ax.plot([1,num_weeks], [args.totalbudget, args.totalbudget],'-k',label='total budget')

    if args.regressionON and week>2:
        ydata = cumsum_billings_by_week[:week]
        xdata = np.arange(1,week+1)
        popt, pcov = curve_fit(linear_func, xdata, ydata)
        ax.plot(np.arange(1,num_weeks+1),linear_func(np.arange(1,num_weeks+1), popt[0], popt[1])/1000,':k',label='linear regression')
        ax.plot(num_weeks,linear_func(num_weeks, popt[0], popt[1])/1000,'ko')
        ax.legend()
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

    ### per employee
    text='Actuals accumulated per week per employee'
    print("generating figure:", text)
    fig, (ax,lax) = plt.subplots(ncols=2, gridspec_kw={"width_ratios":[4,1]})
    bot = np.zeros(num_weeks)
    for a, b in billings_by_employees_by_week.items():
        bc = np.cumsum(b)
        bc[week:] = 0
        ax.bar(np.linspace(1,num_weeks,num_weeks),bc/1000, bottom=bot/1000, label=employees_by_number[a])
        bot+=bc
    ax.set_ylabel('KNOK')
    if args.totalbudget:
        ax.plot([1,num_weeks], [args.totalbudget, args.totalbudget],'-k',label='total budget')

    if args.regressionON and week>2:
        ax.plot(np.arange(1,num_weeks+1),linear_func(np.arange(1,num_weeks+1), popt[0], popt[1])/1000,':k',label='linear regression')
        ax.plot(num_weeks,linear_func(num_weeks, popt[0], popt[1])/1000,'ko')

    h,l = ax.get_legend_handles_labels()
    lax.legend(h,l, borderaxespad=0)
    lax.axis("off")
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

### pie charts
    employeenames = list(employees_by_number.values())
    pie_sizes = np.array(list(billings_by_employees_by_year.values()))
    usedbudget = np.sum(pie_sizes)

    text='Budget actuals'
    print("generating figure:", text)
    fig, ax = plt.subplots()
    ax.pie(pie_sizes/usedbudget, labels=employeenames, autopct='%1.1f%%', shadow=True, startangle=90)
    ax.axis('equal')
    ax.set_title(text)
    plt.tight_layout()
    plt.savefig(text+".png")

    if args.totalbudget:
        text='Budget total'
        print("generating figure:", text)
        explode = np.append(np.zeros_like(pie_sizes), 0.1)
        pie_labels = employeenames.copy()
        pie_labels.append('remaining')
        pie_sizes = np.append(pie_sizes, args.totalbudget*1000-usedbudget)
        pie_sizes = pie_sizes/args.totalbudget*1000
        fig, ax = plt.subplots()
        ax.pie(pie_sizes, explode=explode, labels=pie_labels, autopct='%1.1f%%', shadow=True, startangle=90)
        ax.axis('equal')
        ax.set_title(text)
        plt.tight_layout()
        plt.savefig(text+".png")

    ### print some stats
    ln = len(max(employeenames, key=len))
    tot={}
    for a,b in billings_by_employees_by_year.items():
        tot[a]=str(int(b/1000)).rjust(6, ' ')
    
    print("Billings [KNOK] (modulo round off errors):")
    tmp=str("Employee").ljust(ln, ' ')+" |"
    for i in range(0,12):
        tmp+=labels_month[i].rjust(4, ' ')
    tmp+="| total"
    print(tmp)

    tmp=str("-").ljust(ln, '-')+"--"
    for i in range(0,12):
        tmp+=str("-").rjust(4, '-')
    tmp+="-------"
    print(tmp)

    for a,b in billings_by_employees_by_month.items():
        tmp=employees_by_number[a].ljust(ln, ' ')+" |"
        for i in range(0,12):
            tmp+=str(int(b[i]/1000)).rjust(4, ' ')
        tmp+="|"+tot[a]
        print(tmp)

    tmp=str("-").ljust(ln, '-')+"--"
    for i in range(0,12):
        tmp+=str("-").rjust(4, '-')
    tmp+="-------"
    print(tmp)

    tmp=str("total").ljust(ln, ' ')+" |"
    for i in range(0,12):
        tmp+=str(int(billings_by_month[i]/1000)).rjust(4, ' ')
    tmp+="|"+str(int(usedbudget/1000)).rjust(6,' ')
    print(tmp)
    tmp=str("-").ljust(ln, '-')+"--"
    for i in range(0,12):
        tmp+=str("-").rjust(4, '-')
    tmp+="-------"
    print(tmp)

    if args.totalbudget:
        print("")
        print("Remaining:", args.totalbudget-int(usedbudget/1000), " KNOK")
    print("")

    if not args.totalbudget:
        print("hint: specify total budget with command line option --totalbudget [KNOK]")
